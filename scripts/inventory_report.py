"""Inventory + comparison report between a Telegram source channel and a
forum-supergroup destination across all of its topics.

Indexes both sides into SQLite (resumable), parses movie/series filenames,
classifies each source file by genre, computes match verdicts against every
destination topic, and writes a multi-sheet .xlsx report.

Usage:
    py -3.11 scripts/inventory_report.py \\
        --source -1001416240381 \\
        --dest   -1003776591963

Session resolution (same as downloader.py):
    TELETHON_SESSION_STRING > --session arg > TELETHON_SESSION_FILE env >
    default 'tg_session_acct3'.
"""
from __future__ import annotations

import argparse
import asyncio
import io
import json
import os
import re
import sqlite3
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.tl.functions.messages import GetForumTopicsRequest
from telethon.tl.types import ForumTopic

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent


# ── Filename parser ──────────────────────────────────────────────────────

_RE_YEAR = re.compile(r"\b(19[0-9]\d|20[0-3]\d)\b")
_RE_SE = re.compile(r"\bS(\d{1,2})\s*[._-]?\s*E(\d{1,3})(?:\s*[._-]?\s*E?(\d{1,3}))?\b", re.I)
_RE_SE_ALT = re.compile(r"\bSeason\s*(\d{1,2}).*?Episode\s*(\d{1,3})\b", re.I)
_RE_QUALITY = re.compile(r"\b(2160p|4k|1080p|720p|480p|360p)\b", re.I)
_RE_CODEC = re.compile(r"\b(x265|x264|h\.?265|h\.?264|hevc|avc)\b", re.I)
_RE_SOURCE_TAG = re.compile(
    r"\b(BluRay|BDRip|BRRip|WEB-?DL|WEBRip|HDRip|DVDRip|HDTV|HDCAM|CAMRip|TS|TC|PreDVD)\b", re.I
)
_RE_LANG = re.compile(
    r"\b(English|Hindi|Tamil|Telugu|Malayalam|Kannada|Punjabi|Bengali|Marathi|Gujarati|"
    r"Dual[\s_-]?Audio|Multi(?:[\s_-]?Audio)?)\b", re.I
)
_RE_GROUP = re.compile(r"[-\s]([A-Z][A-Z0-9]{2,8})$")

# Tokens that pollute titles. Stripped after main field extraction.
# Deliberately omits common English words ("the", "of", "in", "from", "and") so
# real movie titles ("The Dark Knight") survive intact.
_TITLE_NOISE_TOKENS = (
    "combined", "complete", "extended", "uncut", "unrated", "remastered",
    "directors", "cut", "director", "theatrical", "imax", "proper", "repack",
    "10bit", "10-bit", "8bit", "esub", "esubs", "msub", "msubs", "subs",
    "ddp5", "ddp2", "ddp", "aac2", "aac", "ac3", "dts", "dd5", "dd2", "dd",
    "atmos", "truehd",
    "hevc", "avc", "x265", "x264", "h265", "h264", "h.265", "h.264", "nvenc",
    "hindi", "english", "dub", "sub", "subbed", "dubbed",
    "season", "episode", "ep",
    "hd", "sd", "uhd", "4kuhd", "fullhd", "hq",
    "yify", "yts", "rarbg", "galaxyrg", "pahe", "psa", "eztv", "anonymous",
    "evo", "exeps", "joy", "jojo", "feliks", "playhd", "tigole", "fitgirl",
    "kogi", "kogimovies", "anoxmous", "mkvcage", "mkvanime",
    "fbm", "amzn", "atvp", "dsnp", "hmax",
    "fzmovies", "hqmovies", "movieshala",
    "brrip", "hdrip", "webrip", "web", "bdrip", "dvdrip", "hdtv", "mkv", "mp4",
    "telly", "ftp", "ddp51", "x2641", "10240k",
)
_RE_EXT = re.compile(r"\.(mkv|mp4|avi|mov|m4v|webm|ts|flv|wmv|mpg|mpeg|3gp|"
                     r"pdf|epub|cbz|cbr|zip|rar|7z)$", re.I)
_RE_SEP = re.compile(r"[._\-\s\(\)\[\]]+")


def _norm_text(s: str) -> str:
    """Lowercase, collapse separators to single spaces, strip."""
    return _RE_SEP.sub(" ", s).strip().lower()


def parse_filename(name: str) -> dict:
    """Extract structured fields from a Telegram-style movie/series filename.

    Returns dict with: title (str), year (int|None), season (int|None),
    ep_lo (int|None), ep_hi (int|None), quality (str|None), codec (str|None),
    source_tag (str|None), language (str|None), logical_key (str), is_series (bool).
    """
    if not name:
        return {
            "title": "", "year": None, "season": None, "ep_lo": None, "ep_hi": None,
            "quality": None, "codec": None, "source_tag": None, "language": None,
            "logical_key": "", "is_series": False,
        }

    work = _RE_EXT.sub("", name)
    work = _RE_GROUP.sub("", work)
    # First pass: normalize `_`, `.`, parens, brackets to spaces so \b fires.
    # Keep `-` for now so multi-token tags like WEB-DL, BD-Rip, x-265, dual-audio
    # still match their `-?` patterns below.
    work = re.sub(r"[._\(\)\[\]]+", " ", work)
    work = re.sub(r"\s+", " ", work).strip()

    se = _RE_SE.search(work) or _RE_SE_ALT.search(work)
    season = ep_lo = ep_hi = None
    if se:
        if se.re is _RE_SE:
            season = int(se.group(1))
            ep_lo = int(se.group(2))
            ep_hi = int(se.group(3)) if se.group(3) else ep_lo
        else:
            season = int(se.group(1))
            ep_lo = ep_hi = int(se.group(2))

    year_m = _RE_YEAR.search(work)
    year = int(year_m.group(1)) if year_m else None

    quality = (_RE_QUALITY.search(work).group(1).lower() if _RE_QUALITY.search(work) else None)
    if quality == "4k":
        quality = "2160p"
    codec_m = _RE_CODEC.search(work)
    codec = codec_m.group(1).lower().replace(".", "") if codec_m else None
    src_m = _RE_SOURCE_TAG.search(work)
    source_tag = src_m.group(1).lower() if src_m else None
    lang_m = _RE_LANG.search(work)
    language = lang_m.group(1).lower().replace("-", "_").replace(" ", "_") if lang_m else None

    # Title extraction strategy: real titles almost always appear BEFORE the
    # year or the S##E## marker. So find the earliest of those two anchors and
    # take everything before it. Falls back to the noise-strip approach if
    # neither anchor exists.
    anchor_idx = None
    for m in (year_m, se):
        if m is None:
            continue
        if anchor_idx is None or m.start() < anchor_idx:
            anchor_idx = m.start()

    if anchor_idx is not None:
        title_work = work[:anchor_idx]
    else:
        title_work = work
        # No year/episode anchor — strip every recognized field instead.
        for pat in (_RE_QUALITY, _RE_CODEC, _RE_SOURCE_TAG, _RE_LANG):
            title_work = pat.sub(" ", title_work)

    # Normalize residual separators (dashes, multi-spaces) and strip leading
    # channel tags (@whatever), bracketed garbage, file-size suffixes, trailing
    # release-group names, etc.
    title_work = title_work.replace("-", " ")
    title_work = re.sub(r"@\w+", " ", title_work)
    title_work = re.sub(r"\b\d+(?:\.\d+)?\s*(?:mb|gb|kb)\b", " ", title_work, flags=re.I)
    title_work = re.sub(r"\b\d+\s*kbps\b", " ", title_work, flags=re.I)
    # Audio channel layouts like 5.1, 7.1, 2.0
    title_work = re.sub(r"\b[257]\.[01]\b", " ", title_work)

    # Drop known noise tokens (whole-word match, case-insensitive).
    for noise in _TITLE_NOISE_TOKENS:
        title_work = re.sub(rf"\b{re.escape(noise)}\b", " ", title_work, flags=re.I)

    # Drop stray ALL-CAPS tokens that are 2–9 chars long — they're almost always
    # release-group abbreviations or audio/source tags (RARBG, YIFY, HW, FBM,
    # NTGFX, ION10, AMZN, etc.). Preserve title-case words. Two false positives
    # we accept: USA/UK/etc. — rare in real movie filenames.
    title_work = re.sub(r"\b[A-Z][A-Z0-9]{1,8}\b", " ", title_work)

    # Drop any token that's a single non-ASCII garbage char or punctuation-only.
    title_work = re.sub(r"[^\w\s]+", " ", title_work, flags=re.UNICODE)

    title = _norm_text(title_work)
    # Strip 1-char dangling tokens at start/end (often "a" or stray letters).
    title = re.sub(r"^\b\w\b\s+", "", title)
    title = re.sub(r"\s+\b\w\b$", "", title).strip()
    title = re.sub(r"\s+", " ", title)

    is_series = season is not None

    if is_series:
        logical_key = f"{title}|s{season:02d}e{ep_lo:02d}-e{ep_hi:02d}"
    elif year:
        logical_key = f"{title}|{year}"
    else:
        logical_key = title

    return {
        "title": title, "year": year, "season": season, "ep_lo": ep_lo, "ep_hi": ep_hi,
        "quality": quality, "codec": codec, "source_tag": source_tag, "language": language,
        "logical_key": logical_key, "is_series": is_series,
    }


# ── Genre classifier ─────────────────────────────────────────────────────

# Topic-title keyword → genre label. Match dest topic titles to these labels.
# Keywords are substring-matched against lowercased topic titles, so a kw of
# "gujrati" matches both "gujrati verse" and "gujarati verse".
_TOPIC_LABEL_KEYWORDS = {
    "SOUTH": ("south",),
    "MARATHI": ("marathi",),
    "GUJARATI": ("gujrati", "gujarati"),
    "ANIME": ("anime",),
    "CARTOON": ("cartoon", "toon"),
    "TV": ("tv show", "tv shows", "tv-show", "tvshow", "series"),
    "OLD_MOVIES": ("old movie", "old-movie", "classic"),
    "THEATRE": ("theatre", "theater", "cam"),
    "DRAMA": ("drama",),
    "MANGA": ("manga", "book", "comic"),
    "COMBINED": ("combines", "combined"),
    "REVIEW": ("review",),
    "MOVIES_AND_SERIES": ("movie & series", "movies & series", "movies and series",
                          "movie and series", "movie&series"),
}

# Heuristics on parsed filename fields → genre label.
_ANIME_TITLES = re.compile(
    r"\b(naruto|one\s*piece|bleach|attack\s*on\s*titan|demon\s*slayer|jujutsu|"
    r"chainsaw\s*man|dragon\s*ball|my\s*hero|hunter\s*x|fullmetal|death\s*note|"
    r"sword\s*art|tokyo\s*ghoul|spy\s*x\s*family|fairy\s*tail|bleach|fma)\b", re.I
)
_CARTOON_TITLES = re.compile(
    r"\b(tom\s*and\s*jerry|spongebob|peppa|doraemon|shinchan|mickey|looney|"
    r"scooby|powerpuff|ben\s*10|tom\s*&\s*jerry|popeye|dexter)\b", re.I
)
_SOUTH_LANGS = {"tamil", "telugu", "malayalam", "kannada"}


def classify_genre(parsed: dict, filename: str) -> tuple[str, str]:
    """Return (genre_label, confidence: 'high'|'med'|'low')."""
    lang = parsed.get("language") or ""
    if lang in _SOUTH_LANGS:
        return "SOUTH", "high"
    if lang == "marathi":
        return "MARATHI", "high"
    if lang == "gujarati":
        return "GUJARATI", "high"
    if _ANIME_TITLES.search(filename or ""):
        return "ANIME", "high"
    if "anime" in (filename or "").lower():
        return "ANIME", "med"
    if _CARTOON_TITLES.search(filename or ""):
        return "CARTOON", "high"
    if "cartoon" in (filename or "").lower() or "kids" in (filename or "").lower():
        return "CARTOON", "med"
    if parsed.get("is_series"):
        return "TV", "med"
    yr = parsed.get("year")
    if yr and yr < 1990:
        return "OLD_MOVIES", "high"
    if (parsed.get("source_tag") or "") in {"camrip", "hdcam", "ts", "tc", "predvd"}:
        return "THEATRE", "high"
    return "MOVIES_AND_SERIES", "low"


def resolve_genre_to_topic(label: str, topics: dict[int, str]) -> tuple[Optional[int], Optional[str]]:
    """Given a genre label and the dest topic map (id→title), return the best
    matching (topic_id, topic_title), or (None, None) if no topic matches."""
    keywords = _TOPIC_LABEL_KEYWORDS.get(label, ())
    best = None
    for tid, title in topics.items():
        tl = title.lower()
        for kw in keywords:
            if kw in tl:
                # Prefer the shortest title (more specific) when multiple match.
                if best is None or len(title) < len(best[1]):
                    best = (tid, title)
    if best:
        return best
    # Fallback: find a "movie...series" or "series...movie" combined topic.
    for tid, title in topics.items():
        tl = title.lower()
        if "movie" in tl and "series" in tl:
            return (tid, title)
    return (None, None)


# ── DB ───────────────────────────────────────────────────────────────────

def setup_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS src_msgs (
            msg_id INTEGER PRIMARY KEY,
            date INTEGER,
            kind TEXT,
            file_name TEXT,
            file_size INTEGER,
            mime TEXT,
            title TEXT,
            year INTEGER,
            season INTEGER,
            ep_lo INTEGER,
            ep_hi INTEGER,
            quality TEXT,
            codec TEXT,
            source_tag TEXT,
            language TEXT,
            logical_key TEXT,
            is_series INTEGER,
            genre TEXT,
            genre_confidence TEXT
        );
        CREATE TABLE IF NOT EXISTS dst_msgs (
            msg_id INTEGER PRIMARY KEY,
            topic_id INTEGER,
            date INTEGER,
            kind TEXT,
            file_name TEXT,
            file_size INTEGER,
            mime TEXT,
            title TEXT,
            year INTEGER,
            season INTEGER,
            ep_lo INTEGER,
            ep_hi INTEGER,
            quality TEXT,
            codec TEXT,
            language TEXT,
            logical_key TEXT
        );
        CREATE TABLE IF NOT EXISTS topics (
            topic_id INTEGER PRIMARY KEY,
            title TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_dst_fnsize ON dst_msgs(file_name, file_size);
        CREATE INDEX IF NOT EXISTS idx_dst_logical ON dst_msgs(logical_key);
        CREATE INDEX IF NOT EXISTS idx_src_logical ON src_msgs(logical_key);
        CREATE INDEX IF NOT EXISTS idx_dst_topic ON dst_msgs(topic_id);
        """
    )
    conn.commit()
    return conn


# ── Message → row helpers ────────────────────────────────────────────────

def _fingerprint(msg) -> tuple[str, Optional[str], int, Optional[str]]:
    """(kind, file_name, file_size_bytes, mime)."""
    doc = getattr(msg, "document", None)
    if doc:
        fn = None
        for a in doc.attributes:
            if hasattr(a, "file_name") and a.file_name:
                fn = a.file_name
                break
        return ("doc", fn, int(doc.size or 0), getattr(doc, "mime_type", None))
    if getattr(msg, "photo", None):
        return ("photo", None, 0, None)
    if getattr(msg, "video", None):
        v = msg.video
        return ("video", None, int(getattr(v, "size", 0) or 0), getattr(v, "mime_type", None))
    if getattr(msg, "message", None):
        return ("text", None, 0, None)
    return ("other", None, 0, None)


def _topic_id_from_message(msg) -> int:
    """Return forum topic id for a dest message. General topic = 1."""
    rt = getattr(msg, "reply_to", None)
    if not rt:
        return 1
    return getattr(rt, "reply_to_top_id", None) or getattr(rt, "reply_to_msg_id", None) or 1


def _src_row(msg, parsed: dict, genre: str, conf: str) -> tuple:
    kind, fn, sz, mime = _fingerprint(msg)
    return (
        msg.id, int(msg.date.timestamp()) if msg.date else None,
        kind, fn, sz, mime,
        parsed["title"], parsed["year"], parsed["season"], parsed["ep_lo"], parsed["ep_hi"],
        parsed["quality"], parsed["codec"], parsed["source_tag"], parsed["language"],
        parsed["logical_key"], int(parsed["is_series"]),
        genre, conf,
    )


def _dst_row(msg, parsed: dict) -> tuple:
    kind, fn, sz, mime = _fingerprint(msg)
    return (
        msg.id, _topic_id_from_message(msg),
        int(msg.date.timestamp()) if msg.date else None,
        kind, fn, sz, mime,
        parsed["title"], parsed["year"], parsed["season"], parsed["ep_lo"], parsed["ep_hi"],
        parsed["quality"], parsed["codec"], parsed["language"],
        parsed["logical_key"],
    )


# ── Indexing ─────────────────────────────────────────────────────────────

async def enumerate_dest_topics(client, dest_id: int, conn: sqlite3.Connection) -> dict[int, str]:
    entity = await client.get_entity(dest_id)
    result = await client(GetForumTopicsRequest(
        peer=entity, q="", offset_date=None, offset_id=0, offset_topic=0, limit=100,
    ))
    topics: dict[int, str] = {}
    rows = []
    for t in result.topics:
        if isinstance(t, ForumTopic):
            topics[t.id] = t.title
            rows.append((t.id, t.title))
    # General topic isn't returned by GetForumTopicsRequest if empty; insert manually.
    if 1 not in topics:
        topics[1] = "General"
        rows.append((1, "General"))
    conn.executemany("INSERT OR REPLACE INTO topics VALUES (?, ?)", rows)
    conn.commit()
    return topics


async def index_source(client, conn: sqlite3.Connection, source_id: int,
                        limit: Optional[int], skip: bool, topics_map: dict[int, str]) -> int:
    if skip:
        n = conn.execute("SELECT COUNT(*) FROM src_msgs").fetchone()[0]
        print(f"[src] skip-index requested; {n:,} rows already in DB", file=sys.stderr)
        return n
    cur = conn.cursor()
    max_indexed = cur.execute("SELECT COALESCE(MAX(msg_id), 0) FROM src_msgs").fetchone()[0]
    kwargs: dict = {"reverse": True}
    if max_indexed > 0:
        kwargs["min_id"] = max_indexed
    if limit:
        kwargs["limit"] = limit
    print(f"[src] indexing from {source_id} (resume past id {max_indexed}, "
          f"limit={limit or 'none'})", file=sys.stderr)
    batch: list[tuple] = []
    count = 0
    started = time.time()
    async for m in client.iter_messages(source_id, **kwargs):
        if getattr(m, "action", None):
            continue  # skip service msgs
        kind, fn, _, _ = _fingerprint(m)
        parsed = parse_filename(fn or "")
        genre, conf = classify_genre(parsed, fn or "")
        batch.append(_src_row(m, parsed, genre, conf))
        count += 1
        if len(batch) >= 500:
            cur.executemany(
                "INSERT OR REPLACE INTO src_msgs VALUES "
                "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
            )
            conn.commit()
            elapsed = time.time() - started
            print(f"[src] {count:,} indexed  ({count/max(elapsed,1):.1f} msg/s)", file=sys.stderr)
            batch.clear()
    if batch:
        cur.executemany(
            "INSERT OR REPLACE INTO src_msgs VALUES "
            "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
        )
        conn.commit()
    total = cur.execute("SELECT COUNT(*) FROM src_msgs").fetchone()[0]
    print(f"[src] done — {count:,} new, {total:,} total", file=sys.stderr)
    return total


async def index_dest(client, conn: sqlite3.Connection, dest_id: int,
                      limit: Optional[int], skip: bool) -> int:
    if skip:
        n = conn.execute("SELECT COUNT(*) FROM dst_msgs").fetchone()[0]
        print(f"[dst] skip-index requested; {n:,} rows already in DB", file=sys.stderr)
        return n
    cur = conn.cursor()
    max_indexed = cur.execute("SELECT COALESCE(MAX(msg_id), 0) FROM dst_msgs").fetchone()[0]
    kwargs: dict = {"reverse": True}
    if max_indexed > 0:
        kwargs["min_id"] = max_indexed
    if limit:
        kwargs["limit"] = limit
    print(f"[dst] indexing from {dest_id} (resume past id {max_indexed}, "
          f"limit={limit or 'none'})", file=sys.stderr)
    batch: list[tuple] = []
    count = 0
    started = time.time()
    async for m in client.iter_messages(dest_id, **kwargs):
        if getattr(m, "action", None):
            continue
        kind, fn, _, _ = _fingerprint(m)
        parsed = parse_filename(fn or "")
        batch.append(_dst_row(m, parsed))
        count += 1
        if len(batch) >= 500:
            cur.executemany(
                "INSERT OR REPLACE INTO dst_msgs VALUES "
                "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
            )
            conn.commit()
            elapsed = time.time() - started
            print(f"[dst] {count:,} indexed  ({count/max(elapsed,1):.1f} msg/s)", file=sys.stderr)
            batch.clear()
    if batch:
        cur.executemany(
            "INSERT OR REPLACE INTO dst_msgs VALUES "
            "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
        )
        conn.commit()
    total = cur.execute("SELECT COUNT(*) FROM dst_msgs").fetchone()[0]
    print(f"[dst] done — {count:,} new, {total:,} total", file=sys.stderr)
    return total


# ── Matching ─────────────────────────────────────────────────────────────

_QUALITY_RANK = {"2160p": 4, "1080p": 3, "720p": 2, "480p": 1, "360p": 0}


def _qual_rank(q: Optional[str]) -> int:
    return _QUALITY_RANK.get(q or "", -1)


def compute_verdicts(conn: sqlite3.Connection) -> dict:
    """Return per-src verdicts + source-duplicate flags. Mutates: writes verdict
    intel back as a temp in-memory dict, keyed by src msg_id.

    For each src doc row, lookup in dst by (file_name, file_size) for exact;
    fallback to logical_key lookup for logical/quality-diff.
    """
    cur = conn.cursor()

    # Build lookup dicts from dst.
    by_fnsize: dict[tuple[str, int], list[tuple[int, int]]] = defaultdict(list)
    by_logical: dict[str, list[tuple[int, int, Optional[str]]]] = defaultdict(list)
    for mid, tid, fn, sz, lk, q in cur.execute(
        "SELECT msg_id, topic_id, file_name, file_size, logical_key, quality "
        "FROM dst_msgs WHERE kind IN ('doc','video') AND file_size > 0 AND file_name IS NOT NULL"
    ):
        if fn:
            by_fnsize[(fn, sz)].append((mid, tid))
        if lk:
            by_logical[lk].append((mid, tid, q))

    # Source duplicates by logical_key (only for docs/videos with parsed key).
    src_by_logical: dict[str, list[tuple[int, Optional[str], int]]] = defaultdict(list)
    for mid, lk, q, sz in cur.execute(
        "SELECT msg_id, logical_key, quality, file_size FROM src_msgs "
        "WHERE kind IN ('doc','video') AND file_size > 0 AND logical_key != ''"
    ):
        if lk:
            src_by_logical[lk].append((mid, q, sz))

    src_dupe_marks: dict[int, str] = {}
    for lk, rows in src_by_logical.items():
        if len(rows) > 1:
            best = max(rows, key=lambda r: (_qual_rank(r[1]), r[2]))
            for mid, q, sz in rows:
                if mid == best[0]:
                    src_dupe_marks[mid] = "BEST"
                else:
                    src_dupe_marks[mid] = "DUPE"

    # Per-src verdict.
    verdicts: dict[int, dict] = {}
    for mid, kind, fn, sz, lk, q in cur.execute(
        "SELECT msg_id, kind, file_name, file_size, logical_key, quality FROM src_msgs"
    ):
        if kind not in ("doc", "video") or not fn or not sz:
            verdicts[mid] = {"verdict": "NOT_A_FILE", "matches_exact": [],
                             "matches_logical": [], "src_dupe": None}
            continue
        exact = by_fnsize.get((fn, sz), [])
        logical = [] if exact else by_logical.get(lk, []) if lk else []
        if exact:
            verdict = "HAVE_EXACT"
        elif logical:
            verdict = "HAVE_LOGICAL_DIFF_QUALITY"
        else:
            verdict = "MISSING"
        verdicts[mid] = {
            "verdict": verdict,
            "matches_exact": exact,
            "matches_logical": logical,
            "src_dupe": src_dupe_marks.get(mid),
        }
    return verdicts


# ── XLSX writer ──────────────────────────────────────────────────────────

def write_xlsx(conn: sqlite3.Connection, verdicts: dict, topics: dict[int, str],
                out_path: Path, source_id: int, dest_id: int) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BOLD = Font(bold=True)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def autosize(ws, max_width=60):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            longest = 0
            for cell in col:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > longest:
                    longest = len(s)
                if longest >= max_width:
                    break
            ws.column_dimensions[letter].width = min(max(10, longest + 2), max_width)

    cur = conn.cursor()
    src_doc_rows = list(cur.execute(
        "SELECT msg_id, date, kind, file_name, file_size, title, year, season, ep_lo, ep_hi, "
        "quality, codec, source_tag, language, logical_key, is_series, genre, genre_confidence "
        "FROM src_msgs ORDER BY msg_id ASC"
    ))

    # ── Sheet 1: Summary ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Inventory report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Source: {source_id}"
    ws["A3"] = f"Destination: {dest_id} (forum, {len(topics)} topics)"
    ws["A4"] = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"].font = Font(italic=True)

    counts = Counter(verdicts[r[0]]["verdict"] for r in src_doc_rows)
    src_total = len(src_doc_rows)
    src_doc_total = sum(1 for r in src_doc_rows if r[2] in ("doc", "video") and r[4])
    dst_doc_total = cur.execute(
        "SELECT COUNT(*) FROM dst_msgs WHERE kind IN ('doc','video') AND file_size > 0"
    ).fetchone()[0]

    row = 6
    ws.cell(row=row, column=1, value="Totals").font = BOLD
    totals_rows = [
        ("Source messages indexed (all kinds)", src_total),
        ("Source files (docs/videos)", src_doc_total),
        ("Destination files (docs/videos, all topics)", dst_doc_total),
        ("", ""),
        ("HAVE_EXACT (same file already in dest)", counts.get("HAVE_EXACT", 0)),
        ("HAVE_LOGICAL_DIFF_QUALITY (same title, diff quality)", counts.get("HAVE_LOGICAL_DIFF_QUALITY", 0)),
        ("MISSING (need to forward)", counts.get("MISSING", 0)),
        ("NOT_A_FILE (text/photo/service)", counts.get("NOT_A_FILE", 0)),
    ]
    for label, val in totals_rows:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)

    # Missing by genre
    row += 3
    ws.cell(row=row, column=1, value="Missing by suggested genre").font = BOLD
    by_genre = Counter()
    by_genre_size = Counter()
    for r in src_doc_rows:
        if verdicts[r[0]]["verdict"] == "MISSING":
            by_genre[r[16]] += 1
            by_genre_size[r[16]] += r[4] or 0
    row += 1
    ws.cell(row=row, column=1, value="Genre").font = BOLD
    ws.cell(row=row, column=2, value="Count").font = BOLD
    ws.cell(row=row, column=3, value="Size (GB)").font = BOLD
    for genre, n in sorted(by_genre.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=genre)
        ws.cell(row=row, column=2, value=n)
        ws.cell(row=row, column=3, value=round(by_genre_size[genre] / 1073741824, 2))

    # Top 20 biggest missing
    row += 3
    ws.cell(row=row, column=1, value="Top 20 biggest MISSING files").font = BOLD
    row += 1
    for h, c in [("file_name", 1), ("size_mb", 2), ("genre", 3)]:
        ws.cell(row=row, column=c, value=h).font = BOLD
    big_missing = sorted(
        [r for r in src_doc_rows if verdicts[r[0]]["verdict"] == "MISSING"],
        key=lambda r: -(r[4] or 0)
    )[:20]
    for r in big_missing:
        row += 1
        ws.cell(row=row, column=1, value=r[3])
        ws.cell(row=row, column=2, value=round((r[4] or 0) / 1048576, 1))
        ws.cell(row=row, column=3, value=r[16])

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    # ── Sheet 2: Source master ─────────────────────────────────────
    ws = wb.create_sheet("Source — master")
    headers = ["src_msg_id", "date", "kind", "file_name", "size_mb",
               "title", "year", "season", "ep_lo", "ep_hi",
               "quality", "codec", "source_tag", "language",
               "is_series", "logical_key",
               "genre", "genre_conf",
               "verdict", "src_dupe",
               "n_matches_exact", "match_topics_exact",
               "n_matches_logical", "match_topics_logical"]
    ws.append(headers)
    for r in src_doc_rows:
        v = verdicts[r[0]]
        exact_topics = ", ".join(sorted({topics.get(t, str(t)) for _, t in v["matches_exact"]}))
        logical_topics = ", ".join(sorted({topics.get(t, str(t)) for _, t, _ in v["matches_logical"]}))
        ws.append([
            r[0],
            time.strftime("%Y-%m-%d", time.localtime(r[1])) if r[1] else "",
            r[2], r[3], round((r[4] or 0) / 1048576, 1),
            r[5], r[6], r[7], r[8], r[9],
            r[10], r[11], r[12], r[13],
            bool(r[15]), r[14],
            r[16], r[17],
            v["verdict"], v["src_dupe"] or "",
            len(v["matches_exact"]), exact_topics,
            len(v["matches_logical"]), logical_topics,
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 3: Missing ───────────────────────────────────────────
    ws = wb.create_sheet("Missing — to forward")
    headers = ["src_msg_id", "file_name", "size_mb", "title", "year",
               "season", "ep_lo", "ep_hi", "quality", "language",
               "genre (suggested)", "genre_conf", "suggested_topic_id", "suggested_topic_title"]
    ws.append(headers)
    missing_rows = [r for r in src_doc_rows if verdicts[r[0]]["verdict"] == "MISSING"]
    missing_rows.sort(key=lambda r: (r[16] or "", -(r[4] or 0)))
    for r in missing_rows:
        tid, ttitle = resolve_genre_to_topic(r[16], topics)
        ws.append([
            r[0], r[3], round((r[4] or 0) / 1048576, 1),
            r[5], r[6], r[7], r[8], r[9],
            r[10], r[13],
            r[16], r[17], tid or "", ttitle or "",
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 4: Already have (exact) ──────────────────────────────
    ws = wb.create_sheet("Already have (exact)")
    headers = ["src_msg_id", "file_name", "size_mb", "title", "year",
               "n_matches", "dst_topic_titles"]
    ws.append(headers)
    for r in src_doc_rows:
        v = verdicts[r[0]]
        if v["verdict"] != "HAVE_EXACT":
            continue
        ws.append([
            r[0], r[3], round((r[4] or 0) / 1048576, 1),
            r[5], r[6],
            len(v["matches_exact"]),
            ", ".join(sorted({topics.get(t, str(t)) for _, t in v["matches_exact"]})),
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 5: Quality differences ───────────────────────────────
    ws = wb.create_sheet("Quality differences")
    headers = ["src_msg_id", "src_filename", "src_quality", "src_size_mb",
               "title", "year", "season", "ep_lo",
               "dst_msg_ids", "dst_qualities", "dst_topic_titles"]
    ws.append(headers)
    for r in src_doc_rows:
        v = verdicts[r[0]]
        if v["verdict"] != "HAVE_LOGICAL_DIFF_QUALITY":
            continue
        dst_ids = ", ".join(str(mid) for mid, _, _ in v["matches_logical"])
        dst_quals = ", ".join((q or "?") for _, _, q in v["matches_logical"])
        dst_topics = ", ".join(sorted({topics.get(t, str(t)) for _, t, _ in v["matches_logical"]}))
        ws.append([
            r[0], r[3], r[10], round((r[4] or 0) / 1048576, 1),
            r[5], r[6], r[7], r[8],
            dst_ids, dst_quals, dst_topics,
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 6: Source duplicates ─────────────────────────────────
    ws = wb.create_sheet("Source duplicates")
    headers = ["logical_key", "src_msg_id", "file_name", "quality", "size_mb", "marker"]
    ws.append(headers)
    src_by_lk: dict[str, list[tuple]] = defaultdict(list)
    for r in src_doc_rows:
        v = verdicts[r[0]]
        if v["src_dupe"]:
            src_by_lk[r[14]].append((r[0], r[3], r[10], r[4] or 0, v["src_dupe"]))
    for lk in sorted(src_by_lk):
        rows = sorted(src_by_lk[lk], key=lambda x: (x[3] == 0, -(x[3] or 0)))
        for mid, fn, q, sz, marker in rows:
            ws.append([lk, mid, fn, q, round((sz or 0) / 1048576, 1), marker])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 7: Destination map ───────────────────────────────────
    ws = wb.create_sheet("Destination map")
    headers = ["dst_msg_id", "topic_id", "topic_title", "file_name", "size_mb",
               "title", "year", "season", "ep_lo", "quality", "language", "logical_key"]
    ws.append(headers)
    for mid, tid, fn, sz, t, y, s, el, eh, q, lng, lk in cur.execute(
        "SELECT msg_id, topic_id, file_name, file_size, title, year, season, ep_lo, ep_hi, "
        "quality, language, logical_key FROM dst_msgs "
        "WHERE kind IN ('doc','video') AND file_size > 0 ORDER BY topic_id, msg_id"
    ):
        ws.append([
            mid, tid, topics.get(tid, ""),
            fn, round((sz or 0) / 1048576, 1),
            t, y, s, el, q, lng, lk,
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 8: Topic stats ───────────────────────────────────────
    ws = wb.create_sheet("Topic stats")
    headers = ["topic_id", "topic_title", "file_count", "total_size_gb", "unique_logical_titles"]
    ws.append(headers)
    rows = cur.execute(
        "SELECT topic_id, COUNT(*), COALESCE(SUM(file_size),0), COUNT(DISTINCT logical_key) "
        "FROM dst_msgs WHERE kind IN ('doc','video') AND file_size > 0 "
        "GROUP BY topic_id ORDER BY 3 DESC"
    ).fetchall()
    for tid, n, total_sz, uniq in rows:
        ws.append([tid, topics.get(tid, ""), n,
                   round(total_sz / 1073741824, 2), uniq])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 9: Genre breakdown ───────────────────────────────────
    ws = wb.create_sheet("Genre breakdown")
    headers = ["genre", "missing_count", "have_count", "missing_size_gb", "coverage_pct"]
    ws.append(headers)
    per_genre_missing: Counter = Counter()
    per_genre_have: Counter = Counter()
    per_genre_missing_size: Counter = Counter()
    for r in src_doc_rows:
        v = verdicts[r[0]]
        if v["verdict"] == "MISSING":
            per_genre_missing[r[16]] += 1
            per_genre_missing_size[r[16]] += r[4] or 0
        elif v["verdict"] in ("HAVE_EXACT", "HAVE_LOGICAL_DIFF_QUALITY"):
            per_genre_have[r[16]] += 1
    all_genres = sorted(set(per_genre_missing) | set(per_genre_have))
    for g in all_genres:
        m, h = per_genre_missing[g], per_genre_have[g]
        cov = round(100 * h / (m + h), 1) if (m + h) else 0
        ws.append([g, m, h, round(per_genre_missing_size[g] / 1073741824, 2), cov])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 10: Unparseable ──────────────────────────────────────
    ws = wb.create_sheet("Unparseable — not a file")
    headers = ["src_msg_id", "kind", "file_name", "size_mb"]
    ws.append(headers)
    for r in src_doc_rows:
        v = verdicts[r[0]]
        if v["verdict"] == "NOT_A_FILE" or (r[2] in ("doc", "video") and not r[5]):
            ws.append([r[0], r[2], r[3] or "", round((r[4] or 0) / 1048576, 1)])
    style_header(ws, len(headers))
    autosize(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[xlsx] wrote {out_path}", file=sys.stderr)


# ── Session loading ──────────────────────────────────────────────────────

def _load_config() -> dict:
    if os.environ.get("TELEGRAM_API_ID") and os.environ.get("TELEGRAM_API_HASH"):
        return {
            "api_id": int(os.environ["TELEGRAM_API_ID"]),
            "api_hash": os.environ["TELEGRAM_API_HASH"],
        }
    return json.loads((ROOT / "config.json").read_text())


def _make_client(session_arg: Optional[str], cfg: dict) -> TelegramClient:
    s = os.environ.get("TELETHON_SESSION_STRING")
    if s:
        return TelegramClient(StringSession(s), cfg["api_id"], cfg["api_hash"])
    name = session_arg or os.environ.get("TELETHON_SESSION_FILE") or "tg_session_acct3"
    if name.endswith(".session"):
        name = name[: -len(".session")]
    return TelegramClient(str(ROOT / name), cfg["api_id"], cfg["api_hash"])


# ── Main ─────────────────────────────────────────────────────────────────

async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=int, required=True)
    ap.add_argument("--dest", type=int, required=True)
    ap.add_argument("--db", default=None)
    ap.add_argument("--report", default=None)
    ap.add_argument("--src-limit", type=int, default=None,
                    help="cap source iteration (testing)")
    ap.add_argument("--dst-limit", type=int, default=None,
                    help="cap dest iteration (testing)")
    ap.add_argument("--skip-src-index", action="store_true")
    ap.add_argument("--skip-dst-index", action="store_true")
    ap.add_argument("--session", default=None,
                    help="session file name (without .session); overrides $TELETHON_SESSION_FILE")
    args = ap.parse_args()

    db_path = Path(args.db) if args.db else (
        ROOT / "data_acct3" / f"inventory_{args.source}.db"
    )
    report_path = Path(args.report) if args.report else (
        ROOT / "data_acct3" / f"inventory_{args.source}.xlsx"
    )

    cfg = _load_config()
    client = _make_client(args.session, cfg)
    await client.connect()
    if not await client.is_user_authorized():
        print("ERROR: session is not authorized. Log in via cli.py first.", file=sys.stderr)
        return 1

    conn = setup_db(db_path)

    try:
        topics = await enumerate_dest_topics(client, args.dest, conn)
        print(f"[topics] dest has {len(topics)} topics:", file=sys.stderr)
        for tid in sorted(topics):
            print(f"    {tid:>5}  {topics[tid]}", file=sys.stderr)

        await index_source(client, conn, args.source, args.src_limit, args.skip_src_index, topics)
        await index_dest(client, conn, args.dest, args.dst_limit, args.skip_dst_index)
    finally:
        await client.disconnect()

    # Re-load topics from DB so the report doesn't need a live client.
    db_topics = dict(conn.execute("SELECT topic_id, title FROM topics").fetchall())

    print(f"[match] computing verdicts...", file=sys.stderr)
    verdicts = compute_verdicts(conn)
    print(f"[match] computed {len(verdicts):,} verdicts", file=sys.stderr)

    write_xlsx(conn, verdicts, db_topics, report_path, args.source, args.dest)
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
