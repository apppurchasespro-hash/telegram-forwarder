"""
Multi-source analyzer: scan N source channels/topics, dedupe against a main
destination channel (and across sources), emit a deterministic forward plan.

READS ONLY — no messages are forwarded by this script.

Usage:
    py -3.11 scripts/multi_source_analyzer.py \\
        --sources-config sources.json \\
        --main -1001234567890 \\
        [--db data_acct3/multi_analysis.db] \\
        [--report data_acct3/multi_analysis.xlsx] \\
        [--plan data_acct3/forward_plan.json] \\
        [--quality-floor 720p] \\
        [--prefer-language hindi,english] \\
        [--max-per-genre 50] \\
        [--main-also-source] \\
        [--topic-genres topic_genres.json] \\
        [--dry-run] \\
        [--src-limit N] [--dst-limit N] \\
        [--skip-src-index] [--skip-dst-index] \\
        [--session NAME]

sources.json format:
    {
      "main": -1001234567890,
      "main_label": "Movies ARENA",
      "sources": [
        {"kind": "topic",   "supergroup": -1003776591963, "topic_id": 12, "label": "Clone/Hindi"},
        {"kind": "topic",   "supergroup": -1003776591963, "topic_id": 18, "label": "Clone/South"},
        {"kind": "channel", "id": -1001416240381, "label": "OldChannelA"}
      ]
    }

topic_genres.json (optional override):
    {"sg:-1003776591963:topic:12": "SOUTH", "sg:-1003776591963:topic:18": "TV"}

Session resolution (same as inventory_report.py):
    TELETHON_SESSION_STRING > --session arg > TELETHON_SESSION_FILE env > tg_session_acct3
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import sqlite3
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

from telethon import TelegramClient
from telethon.errors import FloodWaitError
from telethon.sessions import StringSession
from telethon.tl.functions.messages import GetForumTopicsRequest
from telethon.tl.types import ForumTopic

if hasattr(sys.stdout, "reconfigure") and (sys.stdout.encoding or "").lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure") and (sys.stderr.encoding or "").lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).parent))  # sibling scripts/ importable
from inventory_report import parse_filename, classify_genre, _qual_rank


# ── Fingerprint (4-tuple: kind, file_name, file_size, mime) ─────────────────

def _fingerprint(msg) -> tuple[str, Optional[str], int, Optional[str]]:
    doc = getattr(msg, "document", None)
    if doc:
        fn = None
        for a in doc.attributes:
            if hasattr(a, "file_name") and a.file_name:
                fn = a.file_name
                break
        return ("doc", fn, int(doc.size or 0), getattr(doc, "mime_type", None))
    if getattr(msg, "photo", None):
        return ("photo", None, 0, None)
    vid = getattr(msg, "video", None)
    if vid:
        return ("video", None, int(getattr(vid, "size", 0) or 0), getattr(vid, "mime_type", None))
    if getattr(msg, "message", None):
        return ("text", None, 0, None)
    return ("other", None, 0, None)


# ── Source key ───────────────────────────────────────────────────────────────

def source_key(s: dict) -> str:
    if s["kind"] == "topic":
        return f"sg:{s['supergroup']}:topic:{s['topic_id']}"
    return f"ch:{s['id']}"


# ── DB schema ────────────────────────────────────────────────────────────────

def setup_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sources (
            source_key TEXT PRIMARY KEY,
            label TEXT, kind TEXT,
            supergroup INTEGER, topic_id INTEGER, channel_id INTEGER
        );
        CREATE TABLE IF NOT EXISTS src_msgs (
            source_key TEXT,
            msg_id     INTEGER,
            date INTEGER, kind TEXT, file_name TEXT, file_size INTEGER, mime TEXT,
            title TEXT, year INTEGER, season INTEGER, ep_lo INTEGER, ep_hi INTEGER,
            quality TEXT, codec TEXT, source_tag TEXT, language TEXT,
            logical_key TEXT, is_series INTEGER,
            genre TEXT, genre_confidence TEXT, caption TEXT,
            PRIMARY KEY (source_key, msg_id)
        );
        CREATE TABLE IF NOT EXISTS main_msgs (
            msg_id INTEGER PRIMARY KEY,
            date INTEGER, kind TEXT, file_name TEXT, file_size INTEGER, mime TEXT,
            title TEXT, year INTEGER, season INTEGER, ep_lo INTEGER, ep_hi INTEGER,
            quality TEXT, codec TEXT, language TEXT, logical_key TEXT
        );
        CREATE TABLE IF NOT EXISTS sg_topics (
            supergroup INTEGER, topic_id INTEGER, title TEXT,
            PRIMARY KEY (supergroup, topic_id)
        );
        CREATE TABLE IF NOT EXISTS plan (
            seq INTEGER PRIMARY KEY AUTOINCREMENT,
            source_key TEXT, src_msg_id INTEGER,
            reason TEXT, beaten_main_msg_id INTEGER,
            cluster_id TEXT, rank_in_cluster INTEGER, est_size INTEGER,
            file_name TEXT, title TEXT, year INTEGER,
            season INTEGER, ep_lo INTEGER, ep_hi INTEGER,
            quality TEXT, language TEXT, genre TEXT, clean_caption TEXT
        );
        CREATE TABLE IF NOT EXISTS source_trust (
            source_key TEXT PRIMARY KEY,
            wins INTEGER DEFAULT 0,
            total_clusters INTEGER DEFAULT 0,
            win_rate REAL DEFAULT 0.0
        );
        CREATE INDEX IF NOT EXISTS idx_src_logical ON src_msgs(logical_key);
        CREATE INDEX IF NOT EXISTS idx_src_fnsize  ON src_msgs(file_name, file_size);
        CREATE INDEX IF NOT EXISTS idx_main_logical ON main_msgs(logical_key);
        CREATE INDEX IF NOT EXISTS idx_main_fnsize  ON main_msgs(file_name, file_size);
    """)
    conn.commit()
    return conn


# ── API config + session (mirrors inventory_report._make_client) ─────────────

def _load_api_config() -> dict:
    if os.environ.get("TELEGRAM_API_ID") and os.environ.get("TELEGRAM_API_HASH"):
        return {
            "api_id": int(os.environ["TELEGRAM_API_ID"]),
            "api_hash": os.environ["TELEGRAM_API_HASH"],
        }
    return json.loads((ROOT / "config.json").read_text())


def _make_client(session_arg: Optional[str], cfg: dict) -> TelegramClient:
    s = os.environ.get("TELETHON_SESSION_STRING")
    if s:
        return TelegramClient(StringSession(s), cfg["api_id"], cfg["api_hash"])
    name = session_arg or os.environ.get("TELETHON_SESSION_FILE") or "tg_session_acct3"
    if name.endswith(".session"):
        name = name[: -len(".session")]
    return TelegramClient(str(ROOT / name), cfg["api_id"], cfg["api_hash"])


# ── Enumerate supergroup topics ──────────────────────────────────────────────

async def enumerate_sg_topics(client, sg_id: int, conn: sqlite3.Connection) -> dict[int, str]:
    entity = await client.get_entity(sg_id)
    result = await client(GetForumTopicsRequest(
        peer=entity, q="", offset_date=None, offset_id=0, offset_topic=0, limit=100,
    ))
    topics: dict[int, str] = {}
    rows = []
    for t in result.topics:
        if isinstance(t, ForumTopic):
            topics[t.id] = t.title
            rows.append((sg_id, t.id, t.title))
    if 1 not in topics:
        topics[1] = "General"
        rows.append((sg_id, 1, "General"))
    conn.executemany("INSERT OR REPLACE INTO sg_topics VALUES (?, ?, ?)", rows)
    conn.commit()
    return topics


# ── Source indexing ──────────────────────────────────────────────────────────

def _src_row(sk: str, msg, parsed: dict, genre: str, conf: str, caption: str) -> tuple:
    kind, fn, sz, mime = _fingerprint(msg)
    return (
        sk, msg.id,
        int(msg.date.timestamp()) if msg.date else None,
        kind, fn, sz, mime,
        parsed["title"], parsed["year"], parsed["season"], parsed["ep_lo"], parsed["ep_hi"],
        parsed["quality"], parsed["codec"], parsed["source_tag"], parsed["language"],
        parsed["logical_key"], int(parsed["is_series"]),
        genre, conf, caption,
    )


async def index_one_source(
    client, s: dict, sk: str, conn: sqlite3.Connection,
    limit: Optional[int], skip: bool, topic_genre_override: Optional[str],
) -> int:
    label = s.get("label", sk)
    if skip:
        n = conn.execute("SELECT COUNT(*) FROM src_msgs WHERE source_key=?", (sk,)).fetchone()[0]
        print(f"[src:{label}] skip-index; {n:,} rows already in DB", file=sys.stderr)
        return n

    conn.execute(
        "INSERT OR IGNORE INTO sources VALUES (?,?,?,?,?,?)",
        (sk, label, s["kind"], s.get("supergroup"), s.get("topic_id"), s.get("id"))
    )
    conn.commit()

    cur = conn.cursor()
    max_id = cur.execute(
        "SELECT COALESCE(MAX(msg_id), 0) FROM src_msgs WHERE source_key=?", (sk,)
    ).fetchone()[0]

    iter_kwargs: dict = {"reverse": True}
    if max_id > 0:
        iter_kwargs["min_id"] = max_id
    if limit:
        iter_kwargs["limit"] = limit

    if s["kind"] == "topic":
        iter_kwargs["reply_to"] = s["topic_id"]
        peer = s["supergroup"]
    else:
        peer = s["id"]

    batch: list[tuple] = []
    count = 0
    started = time.time()

    async for m in client.iter_messages(peer, **iter_kwargs):
        if getattr(m, "action", None):
            continue
        kind, fn, sz, mime = _fingerprint(m)
        parsed = parse_filename(fn or "")
        genre, conf = classify_genre(parsed, fn or "")
        if topic_genre_override:
            genre, conf = topic_genre_override, "override"
        caption = (getattr(m, "message", None) or "").strip()
        batch.append(_src_row(sk, m, parsed, genre, conf, caption))
        count += 1
        if len(batch) >= 500:
            cur.executemany(
                "INSERT OR REPLACE INTO src_msgs VALUES "
                "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
            )
            conn.commit()
            elapsed = time.time() - started
            print(f"[src:{label}] {count:,} indexed  ({count/max(elapsed, 1):.0f} msg/s)",
                  file=sys.stderr)
            batch.clear()

    if batch:
        cur.executemany(
            "INSERT OR REPLACE INTO src_msgs VALUES "
            "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
        )
        conn.commit()

    total = cur.execute("SELECT COUNT(*) FROM src_msgs WHERE source_key=?", (sk,)).fetchone()[0]
    print(f"[src:{label}] done — {count:,} new, {total:,} total", file=sys.stderr)
    return total


# ── Main channel indexing ────────────────────────────────────────────────────

def _main_row(msg, parsed: dict) -> tuple:
    kind, fn, sz, mime = _fingerprint(msg)
    return (
        msg.id,
        int(msg.date.timestamp()) if msg.date else None,
        kind, fn, sz, mime,
        parsed["title"], parsed["year"], parsed["season"], parsed["ep_lo"], parsed["ep_hi"],
        parsed["quality"], parsed["codec"], parsed["language"],
        parsed["logical_key"],
    )


async def index_main(
    client, main_id: int, conn: sqlite3.Connection,
    limit: Optional[int], skip: bool,
) -> int:
    if skip:
        n = conn.execute("SELECT COUNT(*) FROM main_msgs").fetchone()[0]
        print(f"[main] skip-index; {n:,} rows already in DB", file=sys.stderr)
        return n

    cur = conn.cursor()
    max_id = cur.execute("SELECT COALESCE(MAX(msg_id), 0) FROM main_msgs").fetchone()[0]
    iter_kwargs: dict = {"reverse": True}
    if max_id > 0:
        iter_kwargs["min_id"] = max_id
    if limit:
        iter_kwargs["limit"] = limit

    batch: list[tuple] = []
    count = 0
    started = time.time()

    async for m in client.iter_messages(main_id, **iter_kwargs):
        if getattr(m, "action", None):
            continue
        kind, fn, sz, mime = _fingerprint(m)
        parsed = parse_filename(fn or "")
        batch.append(_main_row(m, parsed))
        count += 1
        if len(batch) >= 500:
            cur.executemany(
                "INSERT OR REPLACE INTO main_msgs VALUES "
                "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
            )
            conn.commit()
            elapsed = time.time() - started
            print(f"[main] {count:,} indexed  ({count/max(elapsed, 1):.0f} msg/s)",
                  file=sys.stderr)
            batch.clear()

    if batch:
        cur.executemany(
            "INSERT OR REPLACE INTO main_msgs VALUES "
            "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch
        )
        conn.commit()

    total = cur.execute("SELECT COUNT(*) FROM main_msgs").fetchone()[0]
    print(f"[main] done — {count:,} new, {total:,} total", file=sys.stderr)
    return total


# ── Clean caption builder ────────────────────────────────────────────────────

def build_clean_caption(row: dict) -> str:
    """Synthesize a tidy caption from parsed fields (used when source caption is absent/noisy)."""
    lines = []
    title = (row.get("title") or "").title()
    year = row.get("year")
    season, ep_lo, ep_hi = row.get("season"), row.get("ep_lo"), row.get("ep_hi")
    quality = row.get("quality") or ""
    language = (row.get("language") or "").replace("_", " ").title()
    est_size = row.get("est_size") or 0

    title_line = f"\U0001f3ac {title}" if title else ""
    if title_line and year:
        title_line += f" ({year})"
    if title_line:
        lines.append(title_line)

    if season is not None:
        ep_str = f"S{season:02d}E{ep_lo:02d}" if ep_lo else f"S{season:02d}"
        if ep_hi and ep_hi != ep_lo:
            ep_str += f"-E{ep_hi:02d}"
        lines.append(f"\U0001f4fa {ep_str}")

    tags = [t for t in [quality, language] if t]
    if tags:
        lines.append("\U0001f39e " + " · ".join(tags))

    if est_size > 0:
        size_gb = est_size / 1_073_741_824
        size_str = f"{size_gb:.1f} GB" if size_gb >= 1 else f"{est_size / 1_048_576:.0f} MB"
        lines.append(f"\U0001f4e6 {size_str}")

    return "\n".join(lines)


# ── Cluster + plan computation ────────────────────────────────────────────────

def compute_plan(
    conn: sqlite3.Connection,
    quality_floor: Optional[str],
    prefer_langs: list[str],
    max_per_genre: Optional[int],
    main_also_source: bool,
) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """
    Returns (plan_rows, suppressed, upgrades, floor_rejects).
    Clears and repopulates the `plan` and `source_trust` tables.
    """
    cur = conn.cursor()
    floor_rank = _qual_rank(quality_floor) if quality_floor else -2

    # ── Build main lookup
    by_fnsize: dict[tuple, tuple] = {}
    by_logical: dict[str, tuple] = {}
    for mid, fn, sz, lk, q in cur.execute(
        "SELECT msg_id, file_name, file_size, logical_key, quality FROM main_msgs "
        "WHERE kind IN ('doc','video') AND file_size > 0"
    ):
        if fn and sz:
            key = (fn, sz)
            if key not in by_fnsize:
                by_fnsize[key] = (mid, q)
        if lk:
            existing = by_logical.get(lk)
            if existing is None or _qual_rank(q) > _qual_rank(existing[1]):
                by_logical[lk] = (mid, q)

    # ── Load all source rows (docs/videos with size > 0)
    src_rows = list(cur.execute(
        "SELECT source_key, msg_id, date, kind, file_name, file_size, title, year, "
        "season, ep_lo, ep_hi, quality, codec, language, logical_key, is_series, "
        "genre, genre_confidence, caption "
        "FROM src_msgs WHERE kind IN ('doc','video') AND file_size > 0"
    ))

    # Optionally treat main channel as an additional source for cluster ranking
    if main_also_source:
        for mid, date, kind, fn, sz, title, year, season, ep_lo, ep_hi, q, codec, lng, lk in cur.execute(
            "SELECT msg_id, date, kind, file_name, file_size, title, year, season, "
            "ep_lo, ep_hi, quality, codec, language, logical_key "
            "FROM main_msgs WHERE kind IN ('doc','video') AND file_size > 0"
        ):
            src_rows.append(("__main__", mid, date, kind, fn, sz,
                              title, year, season, ep_lo, ep_hi, q, codec, lng, lk,
                              0, "MOVIES_AND_SERIES", "low", ""))

    # ── Cluster by logical_key; apply quality floor
    clusters: dict[str, list] = defaultdict(list)
    floor_rejects: list[dict] = []

    for row in src_rows:
        sk, mid, date, kind, fn, sz, title, year, season, ep_lo, ep_hi, q, codec, lang, lk, is_series, genre, conf, caption = row
        if not fn or not sz:
            continue
        if _qual_rank(q) < floor_rank:
            floor_rejects.append({
                "source_key": sk, "src_msg_id": mid, "file_name": fn,
                "quality": q, "genre": genre,
            })
            continue
        cluster_id = lk if lk else fn
        clusters[cluster_id].append(row)

    # ── Per-cluster verdict
    genre_counts: Counter = Counter()
    plan_rows: list[dict] = []
    suppressed: list[dict] = []
    upgrades: list[dict] = []
    source_wins: Counter = Counter()
    source_totals: Counter = Counter()

    for cluster_id, members in clusters.items():
        if not members:
            continue

        def _rank_key(r):
            sk, mid, date, kind, fn, sz, title, yr, s, el, eh, q, codec, lang, lk, is_s, genre, conf, caption = r
            lang_bonus = 0
            if lang and prefer_langs:
                for pl in prefer_langs:
                    if pl.lower() in (lang or "").lower():
                        lang_bonus = 1
                        break
            return (lang_bonus, _qual_rank(q), sz or 0, -(date or 0))

        members_sorted = sorted(members, key=_rank_key, reverse=True)
        best = members_sorted[0]
        (sk_best, mid_best, date_best, kind_best, fn_best, sz_best, title_best,
         year_best, season_best, ep_lo_best, ep_hi_best, q_best, codec_best,
         lang_best, lk_best, is_series_best, genre_best, conf_best, caption_best) = best

        for r in members:
            source_totals[r[0]] += 1

        # Check against main
        main_exact = by_fnsize.get((fn_best, sz_best)) if (fn_best and sz_best) else None
        main_logical = by_logical.get(lk_best) if lk_best else None

        if main_exact:
            main_mid, main_q = main_exact
            for r in members_sorted:
                suppressed.append({
                    "source_key": r[0], "src_msg_id": r[1], "cluster_id": cluster_id,
                    "reason": "SKIP_HAVE_EXACT", "beaten_by": f"main:{main_mid}",
                    "file_name": r[4], "quality": r[11], "genre": r[16],
                })
            continue

        beaten_main_mid = None
        if main_logical:
            main_mid, main_q = main_logical
            if _qual_rank(q_best) > _qual_rank(main_q):
                reason = "UPGRADE_QUALITY"
                beaten_main_mid = main_mid
            else:
                for r in members_sorted:
                    suppressed.append({
                        "source_key": r[0], "src_msg_id": r[1], "cluster_id": cluster_id,
                        "reason": "SKIP_HAVE_EQUAL_OR_BETTER",
                        "beaten_by": f"main:{main_mid}",
                        "file_name": r[4], "quality": r[11], "genre": r[16],
                    })
                continue
        else:
            reason = "MISSING"

        # Genre throttle
        if max_per_genre and genre_best and genre_counts[genre_best] >= max_per_genre:
            suppressed.append({
                "source_key": sk_best, "src_msg_id": mid_best, "cluster_id": cluster_id,
                "reason": "GENRE_THROTTLED",
                "beaten_by": f"max_per_genre={max_per_genre}",
                "file_name": fn_best, "quality": q_best, "genre": genre_best,
            })
            continue

        # Winner
        source_wins[sk_best] += 1
        r_dict = {
            "source_key": sk_best, "src_msg_id": mid_best,
            "reason": reason,
            "beaten_main_msg_id": beaten_main_mid,
            "cluster_id": cluster_id,
            "rank_in_cluster": 0,
            "est_size": sz_best or 0,
            "file_name": fn_best, "title": title_best, "year": year_best,
            "season": season_best, "ep_lo": ep_lo_best, "ep_hi": ep_hi_best,
            "quality": q_best, "language": lang_best, "genre": genre_best,
        }
        r_dict["clean_caption"] = build_clean_caption(r_dict)
        plan_rows.append(r_dict)
        genre_counts[genre_best or ""] += 1

        if reason == "UPGRADE_QUALITY":
            upgrades.append({**r_dict, "main_msg_id": beaten_main_mid, "main_quality": main_q})

        for r in members_sorted[1:]:
            suppressed.append({
                "source_key": r[0], "src_msg_id": r[1], "cluster_id": cluster_id,
                "reason": "SUPPRESSED_DUPE_CROSS_SOURCE",
                "beaten_by": f"{sk_best}:{mid_best}",
                "file_name": r[4], "quality": r[11], "genre": r[16],
            })

    # Deterministic sort: genre, -year, title, season, ep_lo
    plan_rows.sort(key=lambda r: (
        r.get("genre") or "",
        -(r.get("year") or 0),
        r.get("title") or "",
        r.get("season") or 0,
        r.get("ep_lo") or 0,
    ))

    # Write plan to DB
    conn.execute("DELETE FROM plan")
    conn.executemany(
        "INSERT INTO plan (source_key, src_msg_id, reason, beaten_main_msg_id, "
        "cluster_id, rank_in_cluster, est_size, file_name, title, year, season, "
        "ep_lo, ep_hi, quality, language, genre, clean_caption) VALUES "
        "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(r["source_key"], r["src_msg_id"], r["reason"], r.get("beaten_main_msg_id"),
          r["cluster_id"], r["rank_in_cluster"], r["est_size"],
          r["file_name"], r["title"], r["year"], r["season"], r["ep_lo"], r["ep_hi"],
          r["quality"], r["language"], r["genre"], r["clean_caption"])
         for r in plan_rows]
    )

    # Source trust scores
    all_sks = set(source_wins) | set(source_totals)
    conn.execute("DELETE FROM source_trust")
    conn.executemany(
        "INSERT INTO source_trust VALUES (?,?,?,?)",
        [(sk, source_wins.get(sk, 0), source_totals.get(sk, 0),
          round(source_wins.get(sk, 0) / max(source_totals.get(sk, 1), 1), 3))
         for sk in all_sks]
    )
    conn.commit()

    n_missing = sum(1 for r in plan_rows if r["reason"] == "MISSING")
    n_upgrade = sum(1 for r in plan_rows if r["reason"] == "UPGRADE_QUALITY")
    print(
        f"[plan] {len(plan_rows):,} actions  "
        f"({n_missing} missing, {n_upgrade} upgrades, "
        f"{len(suppressed)} suppressed, {len(floor_rejects)} floor-rejects)",
        file=sys.stderr,
    )
    return plan_rows, suppressed, upgrades, floor_rejects


# ── Plan JSON ────────────────────────────────────────────────────────────────

def write_plan_json(plan_rows: list[dict], main_id: int, path: Path) -> None:
    doc = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "main": main_id,
        "total": len(plan_rows),
        "actions": [
            {
                "seq": i + 1,
                "source_key": r["source_key"],
                "src_msg_id": r["src_msg_id"],
                "reason": r["reason"],
                "beaten_main_msg_id": r.get("beaten_main_msg_id"),
                "cluster_id": r["cluster_id"],
                "est_size": r["est_size"],
                "file_name": r["file_name"],
                "genre": r["genre"],
                "clean_caption": r["clean_caption"],
            }
            for i, r in enumerate(plan_rows)
        ],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=2, ensure_ascii=False)
    tmp.replace(path)
    print(f"[plan] wrote {path}  ({len(plan_rows):,} actions)", file=sys.stderr)


# ── XLSX report ──────────────────────────────────────────────────────────────

def write_xlsx(
    conn: sqlite3.Connection,
    plan_rows: list[dict],
    suppressed: list[dict],
    upgrades: list[dict],
    floor_rejects: list[dict],
    path: Path,
    main_id: int,
    source_labels: dict[str, str],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BOLD = Font(bold=True)

    def style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def autosize(ws, max_width: int = 60) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            longest = max(
                (len(str(c.value)) for c in col if c.value is not None), default=8
            )
            ws.column_dimensions[letter].width = min(max(10, longest + 2), max_width)

    cur = conn.cursor()

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Multi-source analyzer report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Main channel: {main_id}"
    ws["A3"] = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(italic=True)

    src_total = cur.execute("SELECT COUNT(*) FROM src_msgs").fetchone()[0]
    main_total = cur.execute(
        "SELECT COUNT(*) FROM main_msgs WHERE kind IN ('doc','video') AND file_size > 0"
    ).fetchone()[0]
    n_missing = sum(1 for r in plan_rows if r["reason"] == "MISSING")
    n_upgrade = sum(1 for r in plan_rows if r["reason"] == "UPGRADE_QUALITY")
    n_skip = sum(1 for r in suppressed if "SKIP" in r["reason"])
    n_dupe = sum(1 for r in suppressed if "DUPE" in r["reason"] or "THROTTLED" in r["reason"])
    est_total_gb = sum(r["est_size"] for r in plan_rows) / 1_073_741_824

    row = 5
    ws.cell(row=row, column=1, value="Stats").font = BOLD
    for label, val in [
        ("Source messages indexed (all sources)", src_total),
        ("Main channel files indexed", main_total),
        ("", ""),
        ("Plan: MISSING (new to main)", n_missing),
        ("Plan: UPGRADE_QUALITY", n_upgrade),
        ("Suppressed: already in main (exact/better)", n_skip),
        ("Suppressed: cross-source dupes / throttled", n_dupe),
        ("Quality floor rejects", len(floor_rejects)),
        ("", ""),
        ("Estimated forward volume (GB)", round(est_total_gb, 2)),
    ]:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)

    row += 3
    ws.cell(row=row, column=1, value="Plan breakdown by genre").font = BOLD
    row += 1
    for h, c in [("Genre", 1), ("Count", 2), ("Est GB", 3)]:
        ws.cell(row=row, column=c, value=h).font = BOLD
    by_genre: Counter = Counter()
    by_genre_sz: Counter = Counter()
    for r in plan_rows:
        g = r.get("genre") or "?"
        by_genre[g] += 1
        by_genre_sz[g] += r.get("est_size") or 0
    for g, n in sorted(by_genre.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=g)
        ws.cell(row=row, column=2, value=n)
        ws.cell(row=row, column=3, value=round(by_genre_sz[g] / 1_073_741_824, 2))
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    # ── Sheet 2: Plan (to forward) ────────────────────────────────
    ws = wb.create_sheet("Plan (to forward)")
    headers = ["seq", "source_key", "source_label", "src_msg_id", "reason",
               "file_name", "title", "year", "season", "ep_lo", "ep_hi",
               "quality", "language", "genre", "est_size_mb", "cluster_id",
               "beaten_main_msg_id", "clean_caption"]
    ws.append(headers)
    for i, r in enumerate(plan_rows, 1):
        ws.append([
            i, r["source_key"], source_labels.get(r["source_key"], r["source_key"]),
            r["src_msg_id"], r["reason"],
            r["file_name"], r["title"], r["year"], r["season"], r["ep_lo"], r["ep_hi"],
            r["quality"], r["language"], r["genre"],
            round((r.get("est_size") or 0) / 1_048_576, 1),
            r["cluster_id"], r.get("beaten_main_msg_id"), r.get("clean_caption"),
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 3: Upgrades ─────────────────────────────────────────
    ws = wb.create_sheet("Upgrades")
    headers = ["source_key", "src_msg_id", "file_name", "new_quality",
               "main_msg_id", "main_quality", "title", "year", "genre"]
    ws.append(headers)
    for r in upgrades:
        ws.append([
            r["source_key"], r["src_msg_id"], r["file_name"], r["quality"],
            r.get("main_msg_id"), r.get("main_quality"),
            r.get("title"), r.get("year"), r.get("genre"),
        ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 4: Cross-source dupes ───────────────────────────────
    ws = wb.create_sheet("Cross-source dupes")
    headers = ["source_key", "src_msg_id", "file_name", "quality",
               "genre", "reason", "beaten_by"]
    ws.append(headers)
    for r in suppressed:
        if "DUPE" in r["reason"] or "THROTTLED" in r["reason"]:
            ws.append([
                r["source_key"], r["src_msg_id"], r.get("file_name"),
                r.get("quality"), r.get("genre"), r["reason"], r.get("beaten_by"),
            ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 5: Skipped (main has equal/better) ──────────────────
    ws = wb.create_sheet("Skipped — already in main")
    headers = ["source_key", "src_msg_id", "file_name", "quality",
               "genre", "reason", "beaten_by"]
    ws.append(headers)
    for r in suppressed:
        if "SKIP" in r["reason"]:
            ws.append([
                r["source_key"], r["src_msg_id"], r.get("file_name"),
                r.get("quality"), r.get("genre"), r["reason"], r.get("beaten_by"),
            ])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 6: Per-source stats ─────────────────────────────────
    ws = wb.create_sheet("Per-source stats")
    headers = ["source_key", "label", "indexed_msgs",
               "wins", "total_clusters", "win_rate"]
    ws.append(headers)
    for sk, label in source_labels.items():
        n_indexed = cur.execute(
            "SELECT COUNT(*) FROM src_msgs WHERE source_key=?", (sk,)
        ).fetchone()[0]
        trust = cur.execute(
            "SELECT wins, total_clusters, win_rate FROM source_trust WHERE source_key=?", (sk,)
        ).fetchone()
        wins, totals, rate = trust if trust else (0, 0, 0.0)
        ws.append([sk, label, n_indexed, wins, totals, rate])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 7: Source trust ─────────────────────────────────────
    ws = wb.create_sheet("Source trust")
    headers = ["source_key", "label", "wins", "total_clusters", "win_rate"]
    ws.append(headers)
    for sk, label in source_labels.items():
        trust = cur.execute(
            "SELECT wins, total_clusters, win_rate FROM source_trust WHERE source_key=?", (sk,)
        ).fetchone()
        wins, totals, rate = trust if trust else (0, 0, 0.0)
        ws.append([sk, label, wins, totals, rate])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 8: Quality floor rejects ───────────────────────────
    ws = wb.create_sheet("Quality floor rejects")
    headers = ["source_key", "src_msg_id", "file_name", "quality", "genre"]
    ws.append(headers)
    for r in floor_rejects:
        ws.append([r["source_key"], r["src_msg_id"], r["file_name"],
                   r.get("quality"), r.get("genre")])
    style_header(ws, len(headers))
    autosize(ws)

    # ── Sheet 9: Unparseable / not a file ─────────────────────────
    ws = wb.create_sheet("Unparseable — not a file")
    headers = ["source_key", "msg_id", "kind", "file_name", "size_mb"]
    ws.append(headers)
    for sk, mid, kind, fn, sz in cur.execute(
        "SELECT source_key, msg_id, kind, file_name, file_size FROM src_msgs "
        "WHERE kind NOT IN ('doc','video') OR file_size = 0 OR file_name IS NULL"
    ):
        ws.append([sk, mid, kind, fn, round((sz or 0) / 1_048_576, 1)])
    style_header(ws, len(headers))
    autosize(ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[xlsx] wrote {path}", file=sys.stderr)


# ── Main ─────────────────────────────────────────────────────────────────────

async def main() -> int:
    ap = argparse.ArgumentParser(
        description="Multi-source analyzer: N sources → main-channel forward plan (read-only)"
    )
    ap.add_argument("--sources-config", required=True,
                    help="Path to sources.json")
    ap.add_argument("--main", type=int, default=None,
                    help="Main channel id (overrides sources.json 'main')")
    ap.add_argument("--db",     default=None, help="SQLite DB path")
    ap.add_argument("--report", default=None, help="XLSX report path")
    ap.add_argument("--plan",   default=None, help="forward_plan.json path")
    ap.add_argument("--src-limit",  type=int, default=None,
                    help="Cap source iteration (testing)")
    ap.add_argument("--dst-limit",  type=int, default=None,
                    help="Cap main iteration (testing)")
    ap.add_argument("--skip-src-index", action="store_true")
    ap.add_argument("--skip-dst-index", action="store_true")
    ap.add_argument("--quality-floor", default=None,
                    help="Minimum quality to include, e.g. 720p")
    ap.add_argument("--prefer-language", default=None,
                    help="Comma-separated preferred languages, e.g. hindi,english")
    ap.add_argument("--max-per-genre", type=int, default=None,
                    help="Max plan actions per genre")
    ap.add_argument("--main-also-source", action="store_true",
                    help="Include main channel in cluster ranking (enables quality upgrades)")
    ap.add_argument("--topic-genres", default=None,
                    help="Path to topic_genres.json override file")
    ap.add_argument("--dry-run", action="store_true",
                    help="Run all phases but suffix output files with .dryrun")
    ap.add_argument("--session", default=None,
                    help="Session file name (without .session); overrides $TELETHON_SESSION_FILE")
    args = ap.parse_args()

    # Load sources config
    src_cfg_path = Path(args.sources_config)
    if not src_cfg_path.exists():
        print(f"ERROR: sources config not found: {src_cfg_path}", file=sys.stderr)
        return 1
    sources_cfg = json.loads(src_cfg_path.read_text(encoding="utf-8"))
    main_id = args.main or sources_cfg.get("main")
    if not main_id:
        print("ERROR: --main or sources.json 'main' key required", file=sys.stderr)
        return 1
    sources = sources_cfg.get("sources", [])
    if not sources:
        print("ERROR: sources.json 'sources' list is empty", file=sys.stderr)
        return 1

    topic_genres: dict[str, str] = {}
    if args.topic_genres and Path(args.topic_genres).exists():
        topic_genres = json.loads(Path(args.topic_genres).read_text(encoding="utf-8"))

    prefer_langs = (
        [l.strip() for l in args.prefer_language.split(",")]
        if args.prefer_language else []
    )

    dry = args.dry_run
    data_dir = ROOT / "data_acct3"
    abs_main = abs(main_id)
    db_path = (
        Path(args.db) if args.db
        else data_dir / f"multi_analysis_{abs_main}{'_dryrun' if dry else ''}.db"
    )
    report_path = (
        Path(args.report) if args.report
        else data_dir / f"multi_analysis_{abs_main}{'_dryrun' if dry else ''}.xlsx"
    )
    plan_path = (
        Path(args.plan) if args.plan
        else data_dir / f"forward_plan_{abs_main}{'_dryrun' if dry else ''}.json"
    )

    if dry:
        print("[dry-run] output paths suffixed _dryrun; no Telegram writes", file=sys.stderr)

    cfg = _load_api_config()
    client = _make_client(args.session, cfg)
    await client.connect()
    if not await client.is_user_authorized():
        print("ERROR: session not authorized. Log in via cli.py first.", file=sys.stderr)
        await client.disconnect()
        return 1

    conn = setup_db(db_path)
    source_labels: dict[str, str] = {}

    try:
        # Phase 1: enumerate topics for any supergroup sources
        sg_topics_cache: dict[int, dict[int, str]] = {}
        for s in sources:
            sk = source_key(s)
            source_labels[sk] = s.get("label", sk)
            if s["kind"] == "topic":
                sg_id = s["supergroup"]
                if sg_id not in sg_topics_cache:
                    print(f"[topics] enumerating supergroup {sg_id}...", file=sys.stderr)
                    try:
                        sg_topics_cache[sg_id] = await enumerate_sg_topics(
                            client, sg_id, conn
                        )
                        print(
                            f"[topics]   found {len(sg_topics_cache[sg_id])} topics",
                            file=sys.stderr,
                        )
                    except Exception as exc:
                        print(
                            f"[topics] WARN: could not enumerate sg {sg_id}: {exc}",
                            file=sys.stderr,
                        )
                        sg_topics_cache[sg_id] = {}
                tid = s.get("topic_id")
                if tid and tid not in sg_topics_cache.get(sg_id, {}):
                    print(
                        f"[topics] WARN: topic_id {tid} not found in sg {sg_id} — "
                        "will attempt anyway", file=sys.stderr,
                    )

        # Phase 2: index each source
        indexing_complete = True
        for s in sources:
            sk = source_key(s)
            tg_override = topic_genres.get(sk)
            while True:
                try:
                    await index_one_source(
                        client, s, sk, conn,
                        limit=args.src_limit,
                        skip=args.skip_src_index,
                        topic_genre_override=tg_override,
                    )
                    break
                except FloodWaitError as exc:
                    wait = exc.seconds + 2
                    print(f"[flood] waiting {wait}s (source {sk})...", file=sys.stderr)
                    await asyncio.sleep(wait)
                except ConnectionError:
                    print(f"[disconnect] connection lost at {sk} — progress saved, re-run to resume",
                          file=sys.stderr)
                    indexing_complete = False
                    break
            if not indexing_complete:
                break

        # Phase 3: index main channel (only if sources finished this run)
        if indexing_complete:
            while True:
                try:
                    await index_main(
                        client, main_id, conn,
                        limit=args.dst_limit,
                        skip=args.skip_dst_index,
                    )
                    break
                except FloodWaitError as exc:
                    wait = exc.seconds + 2
                    print(f"[flood] waiting {wait}s (main)...", file=sys.stderr)
                    await asyncio.sleep(wait)
                except ConnectionError:
                    print("[disconnect] connection lost indexing main — progress saved, re-run to resume",
                          file=sys.stderr)
                    indexing_complete = False
                    break

    finally:
        try:
            await client.disconnect()
        except Exception:
            pass

    if not indexing_complete:
        conn.close()
        print("[partial] indexing incomplete — re-run to resume from saved progress",
              file=sys.stderr)
        return 0

    # Phase 4: cluster + plan
    print("[plan] computing clusters and verdicts...", file=sys.stderr)
    plan_rows, suppressed, upgrades, floor_rejects = compute_plan(
        conn,
        quality_floor=args.quality_floor,
        prefer_langs=prefer_langs,
        max_per_genre=args.max_per_genre,
        main_also_source=args.main_also_source,
    )

    # Phase 5: emit plan JSON
    write_plan_json(plan_rows, main_id, plan_path)

    # Phase 6: XLSX report
    write_xlsx(
        conn, plan_rows, suppressed, upgrades, floor_rejects,
        report_path, main_id, source_labels,
    )

    conn.close()
    print(
        f"[done] {len(plan_rows):,} actions ready\n"
        f"  plan   → {plan_path}\n"
        f"  report → {report_path}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
